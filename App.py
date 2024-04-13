import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from telethon import TelegramClient
from openpyxl import load_workbook
from tkinter import PhotoImage
import os
from openpyxl import Workbook


def filter_func(item, action, channel):
    return item[1] == action and item[2] == channel


class CustomTkinterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.geometry("800x650")

        self.api_id = 26035258
        self.api_hash = '0de12a831f1faa45cb329d9ca51e0125'
        self.client = TelegramClient('Auth_data', api_id=self.api_id, api_hash=self.api_hash)

        self.data_list = None
        self.views_data = {}
        self.forwards_data = {}

        self.act_label = tk.Label(self, text="Введите номер акции: ")
        self.act_label.pack()
        self.act_entry = tk.Entry(self)
        self.act_entry.pack(pady=5)

        self.chanel_label = tk.Label(self, text="Выберите канал: ")
        self.chanel_label.pack()
        self.chanel_combobox = ttk.Combobox(self)
        self.chanel_combobox.pack(pady=5)

        self.cross_image = PhotoImage(file="cross.png")
        self.check_image = PhotoImage(file="check.png")
        self.frame = tk.Frame(self)
        self.frame.pack(pady=5)

        self.import_button = tk.Button(self.frame, text="Import Excel File", command=self.import_excel_file)
        self.import_button.pack(side='left')

        self.import_label = tk.Label(self.frame, image=self.cross_image)
        self.import_label.pack(side='left', padx=5)

        self.start_button = tk.Button(self, text="Start", command=self.start_main)
        self.start_button.pack(pady=5)

        self.scrollbar = tk.Scrollbar(self)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.output_text = tk.Text(self, yscrollcommand=self.scrollbar.set)
        self.output_text.pack(pady=5, fill=tk.BOTH)

        self.scrollbar.config(command=self.output_text.yview)

        self.error_label = tk.Label(self, text="", fg="red")
        self.error_label.pack()

        self.export_button = tk.Button(self, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack()

        with self.client:
            channels = self.client.loop.run_until_complete(self.get_channels())
            self.chanel_combobox['values'] = channels

    def import_excel_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.wb = load_workbook(filename)
            self.ws = self.wb['Лист1']
            self.data_list = [(cell1.value, cell2.value, cell3.value) for cell1, cell2, cell3 in
                              zip(self.ws['F'][1:], self.ws['G'][1:], self.ws['E'][1:])]
            self.import_label.config(image=self.check_image)

    async def get_channels(self):
        dialogs = await self.client.get_dialogs()
        channels = [dialog.title for dialog in dialogs if dialog.is_channel]
        return channels

    async def main(self, links, chanel):
        views = 0
        forwards = 0
        post_s_video = 0
        post_s_photo = 0
        text_post = 0
        dialogs = await self.client.get_dialogs()
        for dialog in dialogs:
            if dialog.title == chanel:
                messages = await self.client.get_messages(dialog, limit=None)
                self.output_text.insert(tk.END, "\n")
        messages_dict = {str(message.id): message for message in messages}
        for link in links:
            parts = link[0].split('/')
            number = '/'.join(parts[4:])
            if "https://t.me/" not in link[0]:
                self.output_text.insert(tk.END, f"Ссылка {link[0]} неверная. Проверьте правильность ссылки.\n")
                self.output_text.insert(tk.END, "\n")
                self.views_data[link[0]] = "ссылка не телеграмм"
                self.forwards_data[link[0]] = "ссылка не телеграмм"
            else:
                message = messages_dict.get(number)
                if message:
                    view = message.views
                    forward = message.forwards
                    self.views_data[link[0]] = view
                    self.forwards_data[link[0]] = forward
                    consistent = message.message[:50]
                    self.output_text.insert(tk.END, "Ссылка на пост: " + link[0] + "\n")
                    self.output_text.insert(tk.END, "Краткое содержимое поста: " + consistent + "\n")
                    self.output_text.insert(tk.END, "Дата публикации: " + str(message.date) + "\n")
                    self.output_text.insert(tk.END, "Просмотры: " + str(view) + "\n")
                    self.output_text.insert(tk.END, "Репосты: " + str(forward) + "\n")
                    if message.video is not None:
                        self.output_text.insert(tk.END, "Тип: Пост с видео\n")
                        post_s_video += 1
                    elif message.photo is not None:
                        self.output_text.insert(tk.END, "Тип: Пост с фото\n")
                        post_s_photo += 1
                    else:
                        self.output_text.insert(tk.END, "Тип: Пост с текстом\n")
                        text_post += 1
                    forwards += forward
                    views += view
                self.output_text.insert(tk.END, "\n")
        self.output_text.insert(tk.END, "Количество постов с видео: " + str(post_s_video) + "\n")
        self.output_text.insert(tk.END, "Количество постов с фото: " + str(post_s_photo) + "\n")
        self.output_text.insert(tk.END, "Количество текстовых постов: " + str(text_post) + "\n")
        self.output_text.insert(tk.END, "Общее количество просмотров по ссылкам: " + str(views) + "\n")
        self.output_text.insert(tk.END, "Общее количество репостов по ссылкам: " + str(forwards) + "\n")

    def start_main(self):
        if self.data_list is None:  # Проверка импорта файла
            messagebox.showerror("Ошибка", "Пожалуйста, сначала импортируйте файл Excel.")
            return
        act = self.act_entry.get()
        chanel = self.chanel_combobox.get()

        if chanel not in self.chanel_combobox['values']:
            messagebox.showerror("Ошибка", "Выбранный канал не найден. Пожалуйста, выберите существующий канал.")
            return

        filtered_list = [item for item in self.data_list if filter_func(item, act, chanel)]
        unfiltered_list = [item for item in self.data_list if
                           not filter_func(item, act, chanel)]
        if not filtered_list:
            self.error_label.config(text="Такой акции нет. Попробуйте еще раз.")
            self.after(1000, lambda: self.error_label.config(text=""))
            return
        with self.client:
            self.client.loop.run_until_complete(self.main(filtered_list, chanel))
        self.output_text.insert(tk.END, "\n")
        # Отображение ссылок, не прошедших фильтрацию
        self.output_text.insert(tk.END, "Ссылки, не прошедшие фильтрацию:\n")
        for item in unfiltered_list:
            if item[0] is not None:
                self.output_text.insert(tk.END, item[0] + "\n")
        self.output_text.insert(tk.END, "======================================================")

    def export_to_excel(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if filename:
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                link = row[5].value
                if link in self.views_data and link in self.forwards_data:
                    ws.cell(row=i, column=8, value=self.views_data[link])
                    ws.cell(row=i, column=9, value=self.forwards_data[link])
            wb.save(filename)
            messagebox.showinfo("Успех", "Файл успешно экспортирован!")


app = CustomTkinterApp()
app.mainloop()
